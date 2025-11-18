"""
Universal Processor Module
Handles merging of Images, Text, Office, and Binary files into a single PDF.
"""

import os
import io
import logging
from pathlib import Path
from typing import List, Tuple, Optional
from datetime import datetime

# PDF Libraries
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.utils import ImageReader
from pypdf import PdfWriter, PdfReader

# Optional: Excel Support
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from config import get_file_category
from core.file_manager import FileManager

logger = logging.getLogger(__name__)

class UniversalProcessor:
    def __init__(self):
        self.file_manager = FileManager()

    def merge_all_to_pdf(self, filepaths: List[str], output_path: str) -> Tuple[bool, Optional[str]]:
        """
        Universal Merging Strategy:
        - PDF: Append pages directly.
        - Image: Draw image onto a new PDF page.
        - Text/Web/Code: Read text and render nicely onto PDF page.
        - Office (Excel): Parse data and render as text/table on PDF page.
        - Binary/Archive: Generate a "File Info" placeholder card page.
        """
        try:
            writer = PdfWriter()
            success_count = 0
            
            for fpath in filepaths:
                category = get_file_category(fpath)
                logger.info(f"Processing {Path(fpath).name} as {category}")

                try:
                    if category == 'document' and fpath.lower().endswith('.pdf'):
                        self._append_pdf(writer, fpath)
                    elif category == 'image':
                        self._image_to_pdf_pages(writer, fpath)
                    elif category == 'text':
                        self._text_to_pdf_pages(writer, fpath)
                    elif category == 'office':
                        self._office_to_pdf_pages(writer, fpath)
                    elif category == 'binary':
                        self._binary_to_pdf_pages(writer, fpath)
                    else:
                        # Fallback for any other format that passed validation
                        self._binary_to_pdf_pages(writer, fpath)
                    
                    success_count += 1
                except Exception as item_error:
                    logger.error(f"Failed to process item {fpath}: {item_error}")
                    # Add an error page so the merge continues but user sees the failure
                    self._create_error_page(writer, fpath, str(item_error))

            # Write final output
            with open(output_path, "wb") as out_f:
                writer.write(out_f)
            
            return True, f"Merged {success_count} files into PDF."

        except Exception as e:
            logger.error(f"Universal merge failed: {e}", exc_info=True)
            return False, str(e)

    def _append_pdf(self, writer: PdfWriter, filepath: str):
        """Append pages from an existing PDF."""
        reader = PdfReader(filepath)
        for page in reader.pages:
            writer.add_page(page)

    def _image_to_pdf_pages(self, writer: PdfWriter, filepath: str):
        """Draw an image onto a PDF page matching its dimensions."""
        packet = io.BytesIO()
        img = ImageReader(filepath)
        iw, ih = img.getSize()
        
        c = canvas.Canvas(packet, pagesize=(iw, ih))
        c.drawImage(img, 0, 0, width=iw, height=ih)
        c.showPage()
        c.save()
        
        packet.seek(0)
        writer.add_page(PdfReader(packet).pages[0])

    def _text_to_pdf_pages(self, writer: PdfWriter, filepath: str):
        """Render text file content (code, html, txt) onto A4 pages."""
        content, _ = self.file_manager.read_file_safe(filepath)
        if not content: return

        packet = io.BytesIO()
        c = canvas.Canvas(packet, pagesize=A4)
        width, height = A4
        
        # Header
        c.setFont("Helvetica-Bold", 12)
        c.drawString(40, height - 40, f"File: {Path(filepath).name}")
        c.setLineWidth(1)
        c.line(40, height - 45, width - 40, height - 45)
        
        # Content
        text_obj = c.beginText(40, height - 60)
        text_obj.setFont("Courier", 10)
        
        lines = content.split('\n')
        current_line = 0
        max_lines = 55 # Approx lines per A4 page at 10pt
        
        for line in lines:
            # Sanitize
            safe_line = line.replace('\r', '').replace('\t', '    ')
            # Wrap long lines
            chars_per_line = 90
            chunks = [safe_line[i:i+chars_per_line] for i in range(0, len(safe_line), chars_per_line)] or [""]
            
            for chunk in chunks:
                text_obj.textLine(chunk)
                current_line += 1
            
            # New Page if full
            if current_line >= max_lines:
                c.drawText(text_obj)
                c.showPage()
                
                # Reset for new page
                c.setFont("Helvetica-Bold", 12)
                c.drawString(40, height - 40, f"File: {Path(filepath).name} (Cont.)")
                c.line(40, height - 45, width - 40, height - 45)
                
                text_obj = c.beginText(40, height - 60)
                text_obj.setFont("Courier", 10)
                current_line = 0
        
        c.drawText(text_obj)
        c.showPage()
        c.save()
        
        packet.seek(0)
        for page in PdfReader(packet).pages:
            writer.add_page(page)

    def _office_to_pdf_pages(self, writer: PdfWriter, filepath: str):
        """Render Excel data as text tables in PDF."""
        if not HAS_OPENPYXL and filepath.endswith(('.xlsx', '.xls')):
            self._create_error_page(writer, filepath, "Missing library: openpyxl. Install to process Excel.")
            return

        if filepath.endswith(('.xlsx', '.xls')):
            try:
                wb = openpyxl.load_workbook(filepath, data_only=True)
                packet = io.BytesIO()
                c = canvas.Canvas(packet, pagesize=A4)
                width, height = A4
                y = height - 40
                
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    # Sheet Header
                    c.setFont("Helvetica-Bold", 14)
                    c.drawString(40, y, f"Sheet: {sheet} ({Path(filepath).name})")
                    y -= 25
                    c.setFont("Helvetica", 8)
                    
                    for row in ws.iter_rows(values_only=True):
                        # Render row data separated by pipes
                        row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                        
                        # Simple truncation to prevent overflow
                        if len(row_text) > 110: 
                            row_text = row_text[:107] + "..."
                            
                        c.drawString(40, y, row_text)
                        y -= 12
                        
                        # Page break
                        if y < 40:
                            c.showPage()
                            y = height - 40
                            c.setFont("Helvetica", 8)
                    
                    # Spacer between sheets
                    y -= 20
                    if y < 60:
                        c.showPage()
                        y = height - 40

                c.save()
                packet.seek(0)
                for page in PdfReader(packet).pages:
                    writer.add_page(page)
            except Exception as e:
                self._create_error_page(writer, filepath, f"Excel processing error: {e}")
        else:
            # Fallback for PPTX/Other office formats not yet implemented
            self._binary_to_pdf_pages(writer, filepath)

    def _binary_to_pdf_pages(self, writer: PdfWriter, filepath: str):
        """Create a visual 'File Info Card' for non-renderable binaries (EXE, ZIP, etc)."""
        packet = io.BytesIO()
        c = canvas.Canvas(packet, pagesize=A4)
        w, h = A4
        
        # Visual Card Design
        c.setStrokeColorRGB(0.3, 0.3, 0.3)
        c.setFillColorRGB(0.95, 0.95, 0.95) # Light gray background
        c.rect(50, h - 300, w - 100, 200, fill=1)
        
        # Text Color
        c.setFillColorRGB(0, 0, 0)
        
        # Title
        c.setFont("Helvetica-Bold", 16)
        c.drawString(70, h - 140, "FILE ATTACHMENT / BINARY")
        
        # File Details
        c.setFont("Helvetica", 12)
        c.drawString(70, h - 180, f"Filename: {Path(filepath).name}")
        
        # Calculate Size
        try:
            size_bytes = os.path.getsize(filepath)
            if size_bytes > 1024 * 1024:
                size_str = f"{size_bytes / (1024*1024):.2f} MB"
            else:
                size_str = f"{size_bytes/1024:.2f} KB"
        except OSError:
            size_str = "Unknown"
            
        c.drawString(70, h - 200, f"Size: {size_str}")
        
        # File Type
        ext = Path(filepath).suffix.upper()
        c.drawString(70, h - 220, f"Format: {ext} File")
        
        # Disclaimer
        c.setFont("Helvetica-Oblique", 10)
        c.setFillColorRGB(0.4, 0.4, 0.4)
        c.drawString(70, h - 260, "Note: This file cannot be rendered visually in PDF.")
        c.drawString(70, h - 275, "It has been included in this collection for reference.")
        
        c.showPage()
        c.save()
        
        packet.seek(0)
        writer.add_page(PdfReader(packet).pages[0])

    def _create_error_page(self, writer: PdfWriter, filepath: str, error: str):
        """Create a red error page in the PDF for failed files."""
        packet = io.BytesIO()
        c = canvas.Canvas(packet, pagesize=A4)
        w, h = A4
        
        c.setFont("Helvetica-Bold", 14)
        c.setFillColorRGB(0.8, 0, 0) # Red
        c.drawString(50, h/2 + 20, f"ERROR PROCESSING: {Path(filepath).name}")
        
        c.setFont("Courier", 10)
        c.setFillColorRGB(0, 0, 0)
        c.drawString(50, h/2, f"Details: {error}")
        
        c.showPage()
        c.save()
        packet.seek(0)
        writer.add_page(PdfReader(packet).pages[0])